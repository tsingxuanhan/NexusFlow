#!/usr/bin/env python3
"""
数据准备：验证 DeepSeek API + 提取表A数据为 prompt 文本
"""
import os, sys, json, time, requests

# ============================================================
# 1. 验证新 DeepSeek API key
# ============================================================
NEW_KEY = 'sk-41c92afd7cb5461b842a9874f7ed1f2c'
print('=' * 60)
print('1. 验证 DeepSeek API（新 key）')
print('=' * 60)
try:
    t0 = time.time()
    r = requests.post('https://api.deepseek.com/v1/chat/completions',
        headers={'Authorization': f'Bearer {NEW_KEY}', 'Content-Type': 'application/json'},
        json={'model': 'deepseek-chat',
              'messages': [{'role': 'user', 'content': 'Reply with exactly: API_OK'}],
              'max_tokens': 10, 'temperature': 0},
        timeout=30)
    dur = time.time() - t0
    print(f'  Status: {r.status_code} ({dur:.1f}s)')
    if r.status_code == 200:
        data = r.json()
        print(f'  Response: {data["choices"][0]["message"]["content"]}')
        print(f'  Model: {data.get("model", "?")}')
        print(f'  Tokens: {data.get("usage", {})}')
        print('  >>> API 可用！')
    else:
        print(f'  Error: {r.text[:300]}')
        sys.exit(1)
except Exception as e:
    print(f'  Exception: {e}')
    sys.exit(1)

# ============================================================
# 2. 读取表A Excel
# ============================================================
print('\n' + '=' * 60)
print('2. 读取表A Excel')
print('=' * 60)

try:
    import openpyxl
except ImportError:
    print('  openpyxl 未安装，尝试 pip install...')
    os.system(f'"{sys.executable}" -m pip install openpyxl -q')
    import openpyxl

XLSX_PATH = r'C:\Users\ASUS\Desktop\表A_历史数据_1980-2020.xlsx'
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
print(f'  Sheet 数量: {len(wb.sheetnames)}')
print(f'  Sheet 名称: {wb.sheetnames}')

# ============================================================
# 3. 匹配关键指标
# ============================================================
print('\n' + '=' * 60)
print('3. 匹配关键指标')
print('=' * 60)

# IMF WEO 标准代码映射（手动指定，确保准确）
INDICATOR_SHEETS = {
    'GDP_growth':      'NGDP_RPCH',      # GDP constant prices % change
    'Inflation_CPI':   'PCPIPCH',         # Inflation avg consumer prices % change
    'Unemployment':   'LUR',              # Unemployment rate %
    'GDP_per_capita':  'NGDPDPC',         # GDP per capita current prices USD
    'Gov_debt':        'GGXWDN_NGDP',     # General government gross debt % of GDP
}

matched = {}
for indicator, sheet_code in INDICATOR_SHEETS.items():
    if sheet_code in wb.sheetnames:
        matched[indicator] = sheet_code
        print(f'  {indicator:20s} -> "{sheet_code}"')
    else:
        print(f'  {indicator:20s} -> NOT FOUND (sheet "{sheet_code}" missing)')

# 检查是否全部匹配
unmatched = [k for k in INDICATOR_SHEETS if k not in matched]
if unmatched:
    print(f'\n  未匹配的指标: {unmatched}')
    print('  所有 sheet 名称供手动匹配:')
    for sn in wb.sheetnames:
        print(f'    - "{sn}"')

# ============================================================
# 4. 提取数据为文本
# ============================================================
print('\n' + '=' * 60)
print('4. 提取数据')
print('=' * 60)

def extract_sheet_data(ws, indicator_name):
    """从 worksheet 提取数据，返回 (years, countries, data_dict)"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], {}

    # 第一行是表头（国家代码），第一列是年份
    header = rows[0]
    # 找到年份列（通常是第一列）
    year_col = 0
    for i, cell in enumerate(header):
        if cell and ('year' in str(cell).lower() or '年' in str(cell) or 'date' in str(cell).lower()):
            year_col = i
            break

    countries = []
    for i, cell in enumerate(header):
        if i == year_col:
            continue
        if cell:
            countries.append((i, str(cell).strip()))

    years = []
    data = {}
    for row in rows[1:]:
        if row[year_col] is None:
            continue
        try:
            year = int(row[year_col])
        except (ValueError, TypeError):
            continue
        years.append(year)
        data[year] = {}
        for col_idx, country_code in countries:
            val = row[col_idx]
            if val is not None and isinstance(val, (int, float)):
                data[year][country_code] = round(float(val), 4)
            else:
                data[year][country_code] = None

    return years, [c[1] for c in countries], data

all_data = {}
for indicator, sheet_name in matched.items():
    ws = wb[sheet_name]
    years, countries, data = extract_sheet_data(ws, indicator)
    all_data[indicator] = {
        'sheet': sheet_name,
        'years': years,
        'countries': countries,
        'data': data
    }
    valid_count = sum(1 for y in data for c in data[y] if data[y][c] is not None)
    total_count = len(years) * len(countries)
    print(f'  {indicator:20s}: {len(years)}年 x {len(countries)}国, 有效{valid_count}/{total_count} ({100*valid_count/max(total_count,1):.0f}%)')

# ============================================================
# 5. 序列化为文本（云端版 + 端侧摘要版）
# ============================================================
print('\n' + '=' * 60)
print('5. 序列化数据为文本')
print('=' * 60)

# 5a. 云端版：全量数据 CSV 格式
cloud_parts = []
for indicator, info in all_data.items():
    countries = info['countries']
    years = info['years']
    data = info['data']

    # CSV header
    lines = [f'### {indicator} (sheet: {info["sheet"]})']
    header = 'Year,' + ','.join(countries)
    lines.append(header)

    for year in years:
        row_vals = []
        for c in countries:
            v = data[year].get(c)
            row_vals.append(f'{v:.2f}' if v is not None else 'NA')
        lines.append(f'{year},' + ','.join(row_vals))

    cloud_parts.append('\n'.join(lines))

cloud_data_text = '\n\n'.join(cloud_parts)
cloud_token_est = len(cloud_data_text) // 3  # 粗估 tokens
print(f'  云端版数据: {len(cloud_data_text)} chars (~{cloud_token_est} tokens)')

# 5b. 端侧版：关键统计摘要
edge_parts = []
for indicator, info in all_data.items():
    countries = info['countries']
    years = info['years']
    data = info['data']

    lines = [f'### {indicator} 摘要']

    # 每5年采样的均值
    for c in countries:
        vals_by_period = {}
        for y in years:
            v = data[y].get(c)
            if v is not None:
                period = (y // 5) * 5
                vals_by_period.setdefault(period, []).append(v)

        period_means = {p: round(sum(vs)/len(vs), 2) for p, vs in sorted(vals_by_period.items())}
        # 只取最近3个5年期 + 2005-2010（金融危机期）
        recent_periods = sorted(period_means.keys())[-4:]
        summary = ', '.join(f'{p}-{p+4}: {period_means[p]}' for p in recent_periods)
        lines.append(f'  {c}: {summary}')

    edge_parts.append('\n'.join(lines))

edge_data_text = '\n\n'.join(edge_parts)
edge_token_est = len(edge_data_text) // 3
print(f'  端侧版数据: {len(edge_data_text)} chars (~{edge_token_est} tokens)')

# ============================================================
# 6. 保存到文件
# ============================================================
OUTPUT_DIR = r'C:\Users\ASUS\WorkBuddy\2026-07-20-19-24-07'

with open(os.path.join(OUTPUT_DIR, 'data_cloud.txt'), 'w', encoding='utf-8') as f:
    f.write(cloud_data_text)

with open(os.path.join(OUTPUT_DIR, 'data_edge.txt'), 'w', encoding='utf-8') as f:
    f.write(edge_data_text)

# 保存元数据
meta = {
    'indicators': {k: {'sheet': v['sheet'], 'years': v['years'], 'countries': v['countries']}
                   for k, v in all_data.items()},
    'cloud_text_chars': len(cloud_data_text),
    'edge_text_chars': len(edge_data_text),
    'deepseek_api_verified': True,
    'deepseek_key_suffix': NEW_KEY[-6:],
}
with open(os.path.join(OUTPUT_DIR, 'data_meta.json'), 'w', encoding='utf-8') as f:
    json.dump(meta, f, ensure_ascii=False, indent=2)

print(f'\n  已保存: data_cloud.txt, data_edge.txt, data_meta.json')
print(f'\n  完成！可以开始写 benchmark 脚本了。')
