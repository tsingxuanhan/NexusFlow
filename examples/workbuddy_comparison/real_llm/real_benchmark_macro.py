#!/usr/bin/env python3
"""
NexusFlow 端云协同 真实 LLM Benchmark
=====================================
宏观经济分析 + 预测任务（单Agent vs 10Agent协作）

端云调度策略（NexusFlow 核心架构验证）：
  云端（DeepSeek API）: 深度推理任务（策略制定、模式发现、假设验证、经济学解释、审查纠错）
  端侧（Ollama 本地） : 标准化任务（背景调研、基础计算、代码验证、归档整理）

数据：DBnomics IMF WEO 2025.4 真实宏观经济数据（20国×5指标×41年）
"""
import json, os, sys, time, requests, traceback
from datetime import datetime

# ============================================================
# 配置
# ============================================================
DEEPSEEK_API_KEY = os.environ.get('DEEPSEEK_API_KEY', 'sk-your-key-here')
DEEPSEEK_URL = 'https://api.deepseek.com/v1/chat/completions'
DEEPSEEK_MODEL = 'deepseek-chat'

OLLAMA_URL = 'http://127.0.0.1:11434/api/chat'
OLLAMA_MODEL = 'qwen3.5:9b'  # 端侧模型

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# 数据加载
# ============================================================
def load_data():
    """加载云端版和端侧版数据"""
    cloud_path = os.path.join(OUTPUT_DIR, 'data_cloud.txt')
    edge_path = os.path.join(OUTPUT_DIR, 'data_edge.txt')
    with open(cloud_path, 'r', encoding='utf-8') as f:
        cloud = f.read()
    with open(edge_path, 'r', encoding='utf-8') as f:
        edge = f.read()
    return cloud, edge

DATA_CLOUD, DATA_EDGE = load_data()

# 加载表B真值（用于最终对比）
def load_truth():
    """加载表B 2021-2025 真值"""
    import openpyxl
    XLSX_B = r'C:\Users\ASUS\Desktop\表B_回测真值.xlsx'
    wb = openpyxl.load_workbook(XLSX_B, data_only=True)

    INDICATOR_SHEETS = {
        'GDP_growth': 'NGDP_RPCH', 'Inflation_CPI': 'PCPIPCH',
        'Unemployment': 'LUR', 'GDP_per_capita': 'NGDPDPC', 'Gov_debt': 'GGXWDN_NGDP',
    }
    truth = {}
    for indicator, sheet_code in INDICATOR_SHEETS.items():
        if sheet_code not in wb.sheetnames:
            continue
        ws = wb[sheet_code]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        countries = [(i, str(c).strip()) for i, c in enumerate(header) if c and i > 0]
        truth[indicator] = {}
        for row in rows[1:]:
            if row[0] is None:
                continue
            try:
                year = int(row[0])
            except (ValueError, TypeError):
                continue
            truth[indicator][year] = {}
            for col_idx, country in countries:
                val = row[col_idx]
                truth[indicator][year][country] = round(float(val), 2) if val and isinstance(val, (int, float)) else None
    return truth

# ============================================================
# LLM 调用（云端 + 端侧）
# ============================================================
def call_deepseek(system_prompt, user_prompt, max_tokens=4096, temperature=0.7):
    """云端：DeepSeek API 调用"""
    payload = json.dumps({
        'model': DEEPSEEK_MODEL,
        'messages': [
            {'role': 'system', 'content': system_prompt},
            {'role': 'user', 'content': user_prompt}
        ],
        'max_tokens': max_tokens,
        'temperature': temperature
    }).encode('utf-8')

    t0 = time.time()
    try:
        r = requests.post(DEEPSEEK_URL, data=payload, headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {DEEPSEEK_API_KEY}'
        }, timeout=120)
        dur = time.time() - t0
        if r.status_code == 200:
            data = r.json()
            return {
                'content': data['choices'][0]['message']['content'],
                'prompt_tokens': data.get('usage', {}).get('prompt_tokens', 0),
                'completion_tokens': data.get('usage', {}).get('completion_tokens', 0),
                'total_tokens': data.get('usage', {}).get('total_tokens', 0),
                'time_sec': round(dur, 2),
                'tier': 'cloud',
                'model': DEEPSEEK_MODEL,
                'error': None
            }
        else:
            return {'content': '', 'prompt_tokens': 0, 'completion_tokens': 0,
                    'total_tokens': 0, 'time_sec': round(dur, 2),
                    'tier': 'cloud', 'model': DEEPSEEK_MODEL, 'error': f'HTTP {r.status_code}: {r.text[:200]}'}
    except Exception as e:
        return {'content': '', 'prompt_tokens': 0, 'completion_tokens': 0,
                'total_tokens': 0, 'time_sec': round(time.time() - t0, 2),
                'tier': 'cloud', 'model': DEEPSEEK_MODEL, 'error': str(e)}

def call_ollama(system_prompt, user_prompt, max_tokens=1500, temperature=0.7):
    """端侧：Ollama 本地模型调用"""
    t0 = time.time()
    try:
        r = requests.post(OLLAMA_URL, json={
            'model': OLLAMA_MODEL,
            'messages': [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': user_prompt}
            ],
            'stream': False,
            'think': False,
            'think': False,
            'options': {'num_predict': max_tokens, 'temperature': temperature}
        }, timeout=300)
        dur = time.time() - t0
        if r.status_code == 200:
            data = r.json()
            content = data['message'].get('content', '')
            thinking = data['message'].get('thinking', '')
            # 如果 content 为空但 thinking 有内容，使用 thinking
            if not content and thinking:
                content = thinking
            eval_count = data.get('eval_count', 0)
            return {
                'content': content,
                'prompt_tokens': data.get('prompt_eval_count', 0),
                'completion_tokens': eval_count,
                'total_tokens': data.get('prompt_eval_count', 0) + eval_count,
                'time_sec': round(dur, 2),
                'tier': 'edge',
                'model': OLLAMA_MODEL,
                'error': None
            }
        else:
            return {'content': '', 'prompt_tokens': 0, 'completion_tokens': 0,
                    'total_tokens': 0, 'time_sec': round(dur, 2),
                    'tier': 'edge', 'model': OLLAMA_MODEL, 'error': f'HTTP {r.status_code}'}
    except Exception as e:
        return {'content': '', 'prompt_tokens': 0, 'completion_tokens': 0,
                'total_tokens': 0, 'time_sec': round(time.time() - t0, 2),
                'tier': 'edge', 'model': OLLAMA_MODEL, 'error': str(e)}

def call_agent(agent_name, user_prompt, context_packet=''):
    """根据 Agent 的端云分配，调用对应模型"""
    agent = NEXUSFLOW_AGENTS[agent_name]
    full_prompt = user_prompt
    if context_packet:
        full_prompt += f'\n\n--- 前序Agent分析摘要 ---\n{context_packet}'

    # 端侧 Agent 用摘要数据，云端 Agent 用全量数据
    if agent['tier'] == 'edge':
        full_prompt = f'数据摘要（端侧轻量版）：\n{DATA_EDGE[:3000]}\n\n{full_prompt}'
    else:
        full_prompt = f'宏观经济数据（全量）：\n{DATA_CLOUD}\n\n{full_prompt}'

    if agent['tier'] == 'cloud':
        return call_deepseek(agent['system'], full_prompt)
    else:
        return call_ollama(agent['system'], full_prompt)

# ============================================================
# NexusFlow 10 Agent 角色定义（宏观经济版 + 端云分配）
# ============================================================
NEXUSFLOW_AGENTS = {
    'Coordinator': {
        'tier': 'cloud',
        'system': '你是NexusFlow多Agent系统的协调者(Coordinator)。你负责统筹宏观经济分析任务、分配工作给专业Agent、综合各方分析得出最终结论。你不直接做数据分析，而是协调、综合、决策。你的输出必须结构清晰、结论明确。',
    },
    'Planner': {
        'tier': 'cloud',
        'system': '你是NexusFlow的策略师(Planner)。你制定宏观经济分析策略：确定分析框架、选择分析方法、规划预测路径。你善于从全局视角设计分析流程，识别关键经济变量和传导机制。',
    },
    'Researcher': {
        'tier': 'edge',
        'system': '你是NexusFlow的研究员(Researcher)。你负责宏观经济背景调研：梳理各国经济历史、重大事件（金融危机、疫情、政策转向）、结构性特征。你的调研为团队提供背景知识支撑。',
    },
    'Miner': {
        'tier': 'cloud',
        'system': '你是NexusFlow的数据矿工(Miner)。你擅长从宏观经济数据中挖掘隐藏模式：周期性规律、转折点信号、跨国传导路径、异常值检测。你用数据驱动的方式发现规律，不凭直觉。',
    },
    'Assayer': {
        'tier': 'cloud',
        'system': '你是NexusFlow的检验师(Assayer)。你负责验证其他Agent提出的假设和发现：检查数据支撑是否充分、逻辑是否自洽、结论是否稳健。你用批判性思维审视分析质量。',
    },
    'Executor': {
        'tier': 'edge',
        'system': '你是NexusFlow的执行者(Executor)。你负责具体的计算任务：统计量计算、排名打分、预测值生成。你的输出必须是数值化的、可验证的。',
    },
    'Caster': {
        'tier': 'edge',
        'system': '你是NexusFlow的投射者(Caster)。你负责代码层面的验证：检查计算逻辑是否正确、数据是否有误、排名是否合理。你用逻辑和数学验证其他Agent的结论。',
    },
    'Artisan': {
        'tier': 'cloud',
        'system': '你是NexusFlow的工匠(Artisan)。你负责经济学解释：将数据发现转化为经济学叙事，解释背后的因果机制、政策含义、结构性因素。你连接数据与理论。',
    },
    'Reviewer': {
        'tier': 'cloud',
        'system': '你是NexusFlow的审查者(Reviewer)。你做最终审查：检查分析完整性、逻辑一致性、结论可靠性。你标注需要修正的问题，给出改进建议。你的审查确保最终输出的质量。',
    },
    'Archivist': {
        'tier': 'edge',
        'system': '你是NexusFlow的归档师(Archivist)。你负责整理和格式化最终输出：将各Agent的分析综合为结构化报告，确保格式统一、信息完整、可追溯。你的输出是团队交付物。',
    },
}

# 自动填充模型名（根据端云分配）
for _ag in NEXUSFLOW_AGENTS.values():
    _ag['model'] = DEEPSEEK_MODEL if _ag['tier'] == 'cloud' else OLLAMA_MODEL

# ============================================================
# L1 任务定义
# ============================================================
L1_TASK_PROMPT = """请完成以下4项宏观经济分析任务：

【任务1：结构性转折点识别】
识别20国在1980-2020年间经历的重大经济转折点（如金融危机、经济衰退、高速增长期结束等）。对每个转折点，说明年份、国家、触发因素、经济影响。

【任务2：三维度排名】
从以下三个维度对20国进行排名（给出前5名和后5名）：
(a) 增长潜力：长期GDP增速趋势
(b) 经济稳定性：通胀和增长的波动率
(c) 危机韧性：危机后恢复速度

【任务3：跨国相关性分析】
找出GDP增速、通胀、失业率等指标之间的跨国传导关系。哪些国家的经济周期高度同步？哪些指标有领先-滞后关系？

【任务4：2021年5指标预测】
基于历史数据模式，预测2021年以下5个指标的全球（20国）中位数趋势：
1. GDP增长率(%) - 给出预测中位数和区间
2. 通胀率CPI(%) - 给出预测中位数和区间
3. 失业率(%) - 给出预测中位数和区间
4. 人均GDP(USD) - 给出增长趋势
5. 政府债务占GDP比重(%) - 给出趋势判断

对GDP增长率和通胀率，请特别关注2020年COVID冲击后的反弹模式。
"""

PREDICTION_FORMAT = """
请在回答的最后，严格按以下JSON格式输出2021年预测（放在```json代码块中）：
```json
{
  "GDP_growth": {"median": 5.5, "low": 3.0, "high": 7.0, "rationale": "COVID后反弹"},
  "Inflation_CPI": {"median": 3.5, "low": 2.0, "high": 5.0, "rationale": "需求恢复+基数效应"},
  "Unemployment": {"median": 6.0, "low": 4.5, "high": 8.0, "rationale": "..."},
  "GDP_per_capita": {"trend": "下降", "magnitude": "-2%", "rationale": "..."},
  "Gov_debt": {"trend": "上升", "magnitude": "+5pp", "rationale": "..."}
}
```
其中 GDP_growth 和 Inflation_CPI 的 median/low/high 是数值（百分比），rationale 是一句话理由。
"""

# ============================================================
# NexusFlow 10Agent 3轮协作流程
# ============================================================
def run_nexusflow():
    """执行 NexusFlow 10Agent 3轮协作"""
    print('\n' + '=' * 70)
    print('  NexusFlow 10Agent 协作模式（端云混合调度）')
    print('=' * 70)

    results = []
    context_packet = ''  # 跨轮次传递的上下文

    # ===== 第1轮：初步分析 =====
    print('\n--- 第1轮：初步分析 ---')
    round1_agents = ['Coordinator', 'Researcher', 'Miner', 'Executor']
    round1_prompts = {
        'Coordinator': L1_TASK_PROMPT + '\n\n请作为协调者，制定分析计划：如何分工完成上述4项任务？哪些是关键分析路径？',
        'Researcher': '请梳理1980-2020年间影响全球经济的重大事件（金融危机、疫情、政策转向等），为团队分析提供历史背景。',
        'Miner': L1_TASK_PROMPT + '\n\n请专注于任务1（转折点识别）和任务3（跨国相关性）。从数据中挖掘模式。',
        'Executor': '请计算20国的关键统计量：各指标的均值、标准差、最大值、最小值、近5年均值。为排名任务提供数据支撑。',
    }

    for agent_name in round1_agents:
        agent = NEXUSFLOW_AGENTS[agent_name]
        print(f'\n  [{agent_name}] ({agent["tier"]}/{agent["model"]}) ...', end='', flush=True)
        result = call_agent(agent_name, round1_prompts[agent_name], context_packet)
        result['agent'] = agent_name
        result['round'] = 1
        result['task_desc'] = round1_prompts[agent_name][:100]
        results.append(result)
        print(f' {result["time_sec"]}s, {result["total_tokens"]}t' + (' [ERROR]' if result['error'] else ''))

        # 更新 context packet（每个 Agent 贡献摘要）
        if result['content'] and not result['error']:
            summary = result['content'][:800]
            context_packet += f'\n[{agent_name} 第1轮]: {summary}\n'

    # ===== 第2轮：深化+验证 =====
    print('\n--- 第2轮：深化分析 + 交叉验证 ---')
    round2_agents = ['Planner', 'Assayer', 'Caster', 'Artisan']
    round2_prompts = {
        'Planner': '基于第1轮分析，请深化策略：如何完成三维度排名（任务2）和预测（任务4）？提出具体分析框架。',
        'Assayer': '请验证第1轮中Miner发现的转折点和跨国相关性。数据支撑是否充分？有哪些需要修正的结论？',
        'Caster': '请验证Executor计算的统计量是否合理。检查：均值和波动率是否符合经济直觉？有无明显计算错误？',
        'Artisan': '请为第1轮发现的经济转折点和跨国传导关系提供经济学解释。COVID冲击后的恢复路径与2008金融危机有何异同？',
    }

    for agent_name in round2_agents:
        agent = NEXUSFLOW_AGENTS[agent_name]
        print(f'\n  [{agent_name}] ({agent["tier"]}/{agent["model"]}) ...', end='', flush=True)
        result = call_agent(agent_name, round2_prompts[agent_name], context_packet)
        result['agent'] = agent_name
        result['round'] = 2
        result['task_desc'] = round2_prompts[agent_name][:100]
        results.append(result)
        print(f' {result["time_sec"]}s, {result["total_tokens"]}t' + (' [ERROR]' if result['error'] else ''))

        if result['content'] and not result['error']:
            summary = result['content'][:800]
            context_packet += f'\n[{agent_name} 第2轮]: {summary}\n'

    # ===== 第3轮：综合+预测 =====
    print('\n--- 第3轮：综合 + 最终预测 ---')
    round3_agents = ['Reviewer', 'Archivist', 'Coordinator']
    round3_prompts = {
        'Reviewer': '请审查前两轮的分析。标注：(1)需要修正的问题 (2)遗漏的分析维度 (3)预测中需要注意的风险。给出改进建议。',
        'Archivist': '请整理前三轮所有分析，输出结构化报告：任务1转折点、任务2排名、任务3相关性、任务4预测。确保格式统一。',
        'Coordinator': '基于全部3轮分析，请输出最终的2021年预测。' + PREDICTION_FORMAT + '\n\n然后简要总结关键发现（不超过500字）。' + L1_TASK_PROMPT,
    }

    for agent_name in round3_agents:
        agent = NEXUSFLOW_AGENTS[agent_name]
        print(f'\n  [{agent_name}] ({agent["tier"]}/{agent["model"]}) ...', end='', flush=True)
        result = call_agent(agent_name, round3_prompts[agent_name], context_packet)
        result['agent'] = agent_name
        result['round'] = 3
        result['task_desc'] = round3_prompts[agent_name][:100]
        results.append(result)
        print(f' {result["time_sec"]}s, {result["total_tokens"]}t' + (' [ERROR]' if result['error'] else ''))

        if result['content'] and not result['error']:
            summary = result['content'][:800]
            context_packet += f'\n[{agent_name} 第3轮]: {summary}\n'

    # 提取最终预测（Coordinator 第3轮输出）
    final_result = next((r for r in reversed(results) if r['agent'] == 'Coordinator' and r['round'] == 3), None)

    return results, final_result, context_packet

# ============================================================
# 单 Agent 对照
# ============================================================
def run_single_agent():
    """单 Agent 一次性完成所有任务"""
    print('\n' + '=' * 70)
    print('  单Agent 对照模式（全部云端）')
    print('=' * 70)

    system = '你是一位资深宏观经济分析师。请基于提供的数据完成分析任务。'
    user = L1_TASK_PROMPT + PREDICTION_FORMAT

    print('\n  [SingleAgent] (cloud/deepseek-chat) ...', end='', flush=True)
    result = call_deepseek(system, user, max_tokens=4096, temperature=0.7)
    result['agent'] = 'SingleAgent'
    result['round'] = 1
    result['task_desc'] = '单Agent一次性完成全部4项任务'
    print(f' {result["time_sec"]}s, {result["total_tokens"]}t' + (' [ERROR]' if result['error'] else ''))

    return result

# ============================================================
# 预测解析
# ============================================================
def parse_predictions(content):
    """从 LLM 输出中解析预测 JSON"""
    if not content:
        return None

    # 找 JSON 代码块
    import re
    json_blocks = re.findall(r'```json\s*(.*?)```', content, re.DOTALL)
    if not json_blocks:
        # 尝试找 { 开头的 JSON
        json_blocks = re.findall(r'\{[^{}]*"GDP_growth"[^{}]*\}', content, re.DOTALL)

    for block in json_blocks:
        try:
            # 清理
            block = block.strip()
            if not block.startswith('{'):
                continue
            data = json.loads(block)
            if 'GDP_growth' in data or 'Inflation_CPI' in data:
                return data
        except json.JSONDecodeError:
            continue

    return None

# ============================================================
# 结果对比
# ============================================================
def compare_with_truth(nf_pred, sa_pred, truth):
    """与表B真值对比"""
    print('\n' + '=' * 70)
    print('  L2 回测对比')
    print('=' * 70)

    # 2021年真值中位数
    truth_2021 = {}
    for indicator in ['GDP_growth', 'Inflation_CPI', 'Unemployment', 'GDP_per_capita', 'Gov_debt']:
        if indicator in truth and 2021 in truth[indicator]:
            vals = [v for v in truth[indicator][2021].values() if v is not None]
            if vals:
                import statistics
                truth_2021[indicator] = round(statistics.median(vals), 2)

    print(f'\n  2021年真值中位数: {truth_2021}')

    comparison = {'truth_2021': truth_2021}
    for label, pred in [('NexusFlow', nf_pred), ('SingleAgent', sa_pred)]:
        if not pred:
            comparison[label] = {'error': '无法解析预测'}
            continue

        result = {}
        for indicator in ['GDP_growth', 'Inflation_CPI']:
            if indicator in pred and isinstance(pred[indicator], dict):
                pred_median = pred[indicator].get('median')
                pred_low = pred[indicator].get('low')
                pred_high = pred[indicator].get('high')
                truth_val = truth_2021.get(indicator)

                if pred_median is not None and truth_val is not None:
                    try:
                        pred_median = float(pred_median)
                        error = abs(pred_median - truth_val)
                        hit = pred_low <= truth_val <= pred_high if (pred_low and pred_high) else False
                        mape = abs(pred_median - truth_val) / max(abs(truth_val), 0.1) * 100
                        result[indicator] = {
                            'pred_median': pred_median,
                            'pred_interval': [pred_low, pred_high],
                            'truth': truth_val,
                            'abs_error': round(error, 2),
                            'hit': hit,
                            'mape_pct': round(mape, 1)
                        }
                    except (ValueError, TypeError):
                        result[indicator] = {'error': f'无法解析预测值: {pred_median}'}

        comparison[label] = result

        if result:
            print(f'\n  {label}:')
            for ind, r in result.items():
                if 'error' not in r:
                    hit_str = 'HIT' if r['hit'] else 'MISS'
                    print(f'    {ind}: pred={r["pred_median"]} (区间{r["pred_interval"]}), truth={r["truth"]}, MAPE={r["mape_pct"]}% [{hit_str}]')

    return comparison

# ============================================================
# 主执行
# ============================================================
def main():
    start_time = time.time()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    print('=' * 70)
    print(f'  NexusFlow 端云协同 真实 LLM Benchmark')
    print(f'  开始时间: {timestamp}')
    print(f'  云端模型: {DEEPSEEK_MODEL} (DeepSeek API)')
    print(f'  端侧模型: {OLLAMA_MODEL} (Ollama 本地)')
    print(f'  数据规模: 20国 x 5指标 x 41年')
    print('=' * 70)

    # 加载真值
    print('\n加载表B真值...')
    truth = load_truth()
    print(f'  真值加载完成: {len(truth)} 个指标')

    # 1. NexusFlow 10Agent 协作
    nf_results, nf_final, nf_context = run_nexusflow()

    # 2. 单Agent 对照
    sa_result = run_single_agent()

    # 3. 解析预测
    print('\n' + '=' * 70)
    print('  解析预测结果')
    print('=' * 70)

    nf_pred = parse_predictions(nf_final['content']) if nf_final else None
    sa_pred = parse_predictions(sa_result['content'])

    print(f'  NexusFlow 预测解析: {"成功" if nf_pred else "失败"}')
    print(f'  SingleAgent 预测解析: {"成功" if sa_pred else "失败"}')

    if nf_pred:
        print(f'  NexusFlow 预测: {json.dumps(nf_pred, ensure_ascii=False, indent=2)[:500]}')
    if sa_pred:
        print(f'  SingleAgent 预测: {json.dumps(sa_pred, ensure_ascii=False, indent=2)[:500]}')

    # 4. 对比真值
    comparison = compare_with_truth(nf_pred, sa_pred, truth)

    # 5. 汇总统计
    total_time = time.time() - start_time
    nf_cloud_calls = sum(1 for r in nf_results if r['tier'] == 'cloud')
    nf_edge_calls = sum(1 for r in nf_results if r['tier'] == 'edge')
    nf_cloud_tokens = sum(r['total_tokens'] for r in nf_results if r['tier'] == 'cloud')
    nf_edge_tokens = sum(r['total_tokens'] for r in nf_results if r['tier'] == 'edge')
    sa_tokens = sa_result['total_tokens']

    summary = {
        'timestamp': timestamp,
        'total_time_sec': round(total_time, 1),
        'nexusflow': {
            'total_calls': len(nf_results),
            'cloud_calls': nf_cloud_calls,
            'edge_calls': nf_edge_calls,
            'cloud_model': DEEPSEEK_MODEL,
            'edge_model': OLLAMA_MODEL,
            'cloud_tokens': nf_cloud_tokens,
            'edge_tokens': nf_edge_tokens,
            'total_tokens': nf_cloud_tokens + nf_edge_tokens,
            'errors': sum(1 for r in nf_results if r['error']),
        },
        'single_agent': {
            'calls': 1,
            'model': DEEPSEEK_MODEL,
            'tokens': sa_tokens,
            'errors': 1 if sa_result['error'] else 0,
        },
        'comparison': comparison,
    }

    print('\n' + '=' * 70)
    print('  执行摘要')
    print('=' * 70)
    print(f'  总时间: {total_time:.0f}秒 ({total_time/60:.1f}分钟)')
    print(f'  NexusFlow: {len(nf_results)}次调用 ({nf_cloud_calls}云+{nf_edge_calls}端), {nf_cloud_tokens+nf_edge_tokens} tokens')
    print(f'  SingleAgent: 1次调用, {sa_tokens} tokens')
    print(f'  错误数: NF={summary["nexusflow"]["errors"]}, SA={summary["single_agent"]["errors"]}')

    # 6. 保存完整结果
    all_results = {
        'summary': summary,
        'nexusflow_results': [{k: v for k, v in r.items() if k != 'content'} for r in nf_results],
        'nexusflow_outputs': {r['agent'] + f'_R{r["round"]}': r['content'][:2000] for r in nf_results},
        'nexusflow_final_output': nf_final['content'] if nf_final else '',
        'nexusflow_context_packet': nf_context,
        'single_agent_result': {k: v for k, v in sa_result.items() if k != 'content'},
        'single_agent_output': sa_result['content'],
        'nf_predictions': nf_pred,
        'sa_predictions': sa_pred,
        'comparison': comparison,
    }

    results_path = os.path.join(OUTPUT_DIR, 'real_benchmark_results.json')
    with open(results_path, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    print(f'\n  完整结果已保存: {results_path}')

    return all_results

if __name__ == '__main__':
    main()
